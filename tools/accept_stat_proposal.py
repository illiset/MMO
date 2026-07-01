"""Apply the accepted stat-minimums proposal (2026-07-01), in the
spreadsheet's own idiom: class minimums are formulas over archetype floors.

- Every x-mark in the stats section becomes `=SUM(<floor cell>+2)` — the
  same rule the Great North classes already encode.
- Exception (the "Volva idiom", per Daniel's decision): DPS-archetype
  classes keying Logic reference the DPS *Knowledge* floor instead, so
  caster DPS need Logic 9 while the archetype-wide Logic floor stays 1
  (keeping each archetype's floors summing to exactly 100).
- Floors are not modified.

Updates BOTH the repo snapshot and the authoritative Downloads master.
Afterwards: recalc (LibreOffice), then re-run extract_from_xlsx.py.
"""
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parent.parent
TARGETS = [
    REPO / "design" / "MMOStats-source-snapshot.xlsx",
    Path(r"C:\Users\delli\Downloads\MMOStats.xlsx"),
]
ARCHETYPE_NAMES = {"Frontline", "Healers", "Support", "DPS"}


def cell_str(v):
    return str(v).strip() if v is not None else ""


def apply(path):
    wb = openpyxl.load_workbook(path)
    total, crossrefs = 0, 0
    for ws in wb.worksheets:
        arch_col = None
        col_to_arch = {}
        dps_col = None
        for col in range(6, 41):
            name = cell_str(ws.cell(1, col).value)
            if name in ARCHETYPE_NAMES:
                arch_col = col
                if name == "DPS":
                    dps_col = col
            elif name and arch_col:
                col_to_arch[col] = arch_col

        stats_row = next(r for r in range(1, 61)
                         if cell_str(ws.cell(r, 4).value) == "Stats")
        knowledge_row = next(r for r in range(stats_row, 61)
                             if cell_str(ws.cell(r, 5).value) == "Knowledge")
        for r in range(stats_row, 61):
            stat_name = cell_str(ws.cell(r, 5).value)
            if not stat_name:
                continue
            for col, acol in col_to_arch.items():
                if cell_str(ws.cell(r, col).value).lower() != "x":
                    continue
                ref_row = r
                if stat_name == "Logic" and acol == dps_col:
                    ref_row = knowledge_row
                    crossrefs += 1
                ws.cell(r, col).value = f"=SUM({get_column_letter(acol)}{ref_row}+2)"
                total += 1
    wb.save(path)
    print(f"{path.name}: {total} formulas written ({crossrefs} Volva-idiom Logic crossrefs)")


for t in TARGETS:
    apply(t)
