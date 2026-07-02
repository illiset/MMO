"""Export all Great North skills to an Excel review workbook for Daniel.

He edits Name/Cost/Cooldown/BaseValue/Description (and NewName/Notes) in
Excel; a future ingest tool reads the edits back into the JSON and flips
status draft->authored. Yellow status cells = drafts needing review.
"""
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parent.parent
SRC = REPO / "data" / "skills" / "great-north"
OUT = REPO / "design" / "skills-review-great-north.xlsx"

HEADERS = ["Tree", "Class", "Class Type", "Skill Name", "Tier", "Level",
           "Kind", "Cost Type", "Cost", "Cooldown (s)", "Base Value",
           "Description", "Status", "Your New Name", "Your Notes"]
DRAFT_FILL = PatternFill("solid", start_color="FFF2AB")
AUTHORED_FILL = PatternFill("solid", start_color="D6E8D0")

wb = Workbook()
ws = wb.active
ws.title = "GN Skills"
ws.append(HEADERS)
for cell in ws[1]:
    cell.font = Font(bold=True, name="Arial")

def add_rows(data, tree_label, class_name, class_type):
    for s in data["skills"]:
        ws.append([tree_label, class_name, class_type, s["name"], s["tier"],
                   s["levelReq"], s["kind"], s["costType"], s["cost"],
                   s["cooldown"], s["baseValue"], s["description"],
                   s["status"], "", ""])
        status_cell = ws.cell(ws.max_row, 13)
        status_cell.fill = DRAFT_FILL if s["status"] == "draft" else AUTHORED_FILL

for f in sorted((SRC / "archetypes").glob("*.json")):
    d = json.loads(f.read_text(encoding="utf-8"))
    add_rows(d, "class-type (1-9)", "-", d["classType"])
for f in sorted((SRC / "classes").glob("*.json")):
    d = json.loads(f.read_text(encoding="utf-8"))
    add_rows(d, "class (10+)", d["className"], d["classType"])

widths = [14, 14, 12, 26, 5, 6, 9, 9, 6, 12, 10, 62, 10, 24, 30]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.font = Font(name="Arial", size=10)

wb.save(OUT)
print(f"wrote {OUT.name}: {ws.max_row - 1} skills")
