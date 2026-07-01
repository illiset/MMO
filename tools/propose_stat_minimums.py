"""Generate proposed numeric class stat minimums for every class that only
has x-marks, using the empirical Great North rule: class min = archetype
floor + 2 (capped at 10).

Reads the committed snapshot, writes a review copy to
design/MMOStats-proposed-stat-minimums.xlsx with proposed cells highlighted
(yellow fill, blue bold text). Does NOT touch the canonical snapshot or data/.
"""
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

REPO = Path(__file__).resolve().parent.parent
SRC = REPO / "design" / "MMOStats-source-snapshot.xlsx"
OUT = REPO / "design" / "MMOStats-proposed-stat-minimums.xlsx"
ARCHETYPE_NAMES = {"Frontline", "Healers", "Support", "DPS"}

PROPOSED_FILL = PatternFill("solid", start_color="FFFF00")
PROPOSED_FONT = Font(bold=True, color="0000FF")


def cell_str(v):
    return str(v).strip() if v is not None else ""


def main():
    wb = openpyxl.load_workbook(SRC)
    total, flags = 0, []
    for ws in wb.worksheets:
        # Map each class column to its archetype column (row 1, left to right).
        arch_col = None
        col_to_arch = {}
        for col in range(6, 41):
            name = cell_str(ws.cell(1, col).value)
            if name in ARCHETYPE_NAMES:
                arch_col = col
            elif name and arch_col:
                col_to_arch[col] = arch_col

        stats_row = next(r for r in range(1, 61)
                         if cell_str(ws.cell(r, 4).value) == "Stats")
        for r in range(stats_row, 61):
            stat_name = cell_str(ws.cell(r, 5).value)
            if not stat_name:
                continue
            for col, acol in col_to_arch.items():
                if cell_str(ws.cell(r, col).value).lower() != "x":
                    continue
                floor = ws.cell(r, acol).value
                if not isinstance(floor, (int, float)):
                    flags.append(f"{ws.title}/{stat_name}: x at col {col} but no "
                                 f"archetype floor — left as x")
                    continue
                proposed = min(int(floor) + 2, 10)
                cell = ws.cell(r, col)
                cell.value = proposed
                cell.fill = PROPOSED_FILL
                cell.font = PROPOSED_FONT
                total += 1
                if floor <= 2:
                    flags.append(f"{ws.title}/{stat_name} "
                                 f"({cell_str(ws.cell(1, col).value)}): floor is "
                                 f"{int(floor)} so proposed {proposed} — floor "
                                 f"looks like a placeholder, review")
    wb.save(OUT)
    print(f"proposed {total} stat minimums -> {OUT.name}")
    for f in flags:
        print("FLAG", f)


if __name__ == "__main__":
    main()
